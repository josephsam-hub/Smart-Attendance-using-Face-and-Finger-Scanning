import tkinter as tk
from tkinter import ttk, messagebox
import cv2
import os
import dlib
import numpy as np
import face_recognition
from pyzbar.pyzbar import decode
from datetime import datetime
import pandas as pd
from datetime import datetime
import sqlite3
from PIL import Image, ImageTk
import hashlib
import winsound  # For beep on Windows
import openpyxl  # For editing Excel files

# Paths and Models
DATASET_DIR = "face_dataset"
CSV_FILENAME = "attendance.csv"
EXCEL_FILENAME = "attendance.xlsx"
SHAPE_PREDICTOR_PATH = "shape_predictor_68_face_landmarks.dat"
FACE_RECOGNITION_MODEL_PATH = "dlib_face_recognition_resnet_model_v1.dat"

# Dlib models
face_detector = dlib.get_frontal_face_detector()
landmark_predictor = dlib.shape_predictor(SHAPE_PREDICTOR_PATH)
face_rec_model = dlib.face_recognition_model_v1(FACE_RECOGNITION_MODEL_PATH)

# SQLite setup
conn = sqlite3.connect("users.db")
c = conn.cursor()
c.execute("""
CREATE TABLE IF NOT EXISTS users (
    username TEXT PRIMARY KEY,
    password TEXT NOT NULL
)
""")
conn.commit()

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

# GUI App
class FaceAttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Face Recognition Attendance System")
        self.root.geometry("900x650")

        self.already_marked_today = set()

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=1, fill='both')

        self.login_tab()
        self.signup_tab()
        self.register_tab()
        self.recognition_tab()

    def login_tab(self):
        self.login_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.login_frame, text="Login")

        ttk.Label(self.login_frame, text="Username").pack(pady=10)
        self.login_username = ttk.Entry(self.login_frame)
        self.login_username.pack()

        ttk.Label(self.login_frame, text="Password").pack(pady=10)
        self.login_password = ttk.Entry(self.login_frame, show="*")
        self.login_password.pack()

        ttk.Button(self.login_frame, text="Login", command=self.login_user).pack(pady=20)

    def signup_tab(self):
        self.signup_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.signup_frame, text="Signup")

        ttk.Label(self.signup_frame, text="Username").pack(pady=10)
        self.signup_username = ttk.Entry(self.signup_frame)
        self.signup_username.pack()

        ttk.Label(self.signup_frame, text="Password").pack(pady=10)
        self.signup_password = ttk.Entry(self.signup_frame, show="*")
        self.signup_password.pack()

        ttk.Button(self.signup_frame, text="Create Account", command=self.signup_user).pack(pady=20)

    def register_tab(self):
        self.register_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.register_frame, text="Register Face")

        ttk.Label(self.register_frame, text="Student Name").pack(pady=5)
        self.name_entry = ttk.Entry(self.register_frame)
        self.name_entry.pack()

        ttk.Label(self.register_frame, text="Roll Number").pack(pady=5)
        self.roll_entry = ttk.Entry(self.register_frame)
        self.roll_entry.pack()

        ttk.Label(self.register_frame, text="Department").pack(pady=5)
        self.dept_entry = ttk.Entry(self.register_frame)
        self.dept_entry.pack()

        ttk.Button(self.register_frame, text="Register Face", command=self.capture_faces).pack(pady=20)

    def recognition_tab(self):
        self.recognize_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.recognize_frame, text="Recognize & Mark")

        self.feedback_label = ttk.Label(self.recognize_frame, text="", font=("Helvetica", 14))
        self.feedback_label.pack(pady=10)

        ttk.Button(self.recognize_frame, text="Start Recognition", command=self.recognize_and_mark).pack(pady=20)

        # Add Mock Test Button
        # ttk.Button(self.recognize_frame, text="Test Marking (Mock)", command=self.test_mark_attendance_mock).pack(pady=10)

    def login_user(self):
        username = self.login_username.get()
        password = self.login_password.get()
        hashed_pw = hash_password(password)
        c.execute("SELECT * FROM users WHERE username=? AND password=?", (username, hashed_pw))
        result = c.fetchone()
        if result:
            messagebox.showinfo("Login", "Login successful!")
        else:
            messagebox.showerror("Login", "Invalid credentials")

    def signup_user(self):
        username = self.signup_username.get()
        password = self.signup_password.get()
        if not username or not password:
            messagebox.showerror("Error", "Username and Password required")
            return
        try:
            hashed_pw = hash_password(password)
            c.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, hashed_pw))
            conn.commit()
            messagebox.showinfo("Success", "Account created successfully!")
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "Username already exists")

    def capture_faces(self):
        name = self.name_entry.get()
        roll = self.roll_entry.get()
        dept = self.dept_entry.get()

        if not name or not roll or not dept:
            messagebox.showerror("Error", "Please fill all fields")
            return

        identifier = f"{roll}_{name}_{dept}"
        student_dir = os.path.join(DATASET_DIR, identifier)
        if not os.path.exists(student_dir):
            os.makedirs(student_dir)

        cap = cv2.VideoCapture(0)
        count = 0
        while count < 25:
            ret, frame = cap.read()
            if not ret:
                break
            if count % 5 == 0:
                img_path = os.path.join(student_dir, f"{identifier}_{count}.jpg")
                cv2.imwrite(img_path, frame)
            count += 1
            cv2.imshow("Registering Face - Press Q to Exit", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        cap.release()
        cv2.destroyAllWindows()
        messagebox.showinfo("Done", f"{count} face images captured for {name}")

    def load_registered_faces(self):
        names = []
        encodings = []
        if not os.path.exists(DATASET_DIR):
            os.makedirs(DATASET_DIR)

        for folder in os.listdir(DATASET_DIR):
            folder_path = os.path.join(DATASET_DIR, folder)
            for img_file in os.listdir(folder_path):
                img_path = os.path.join(folder_path, img_file)
                img = cv2.imread(img_path)
                gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                faces = face_detector(gray)
                if faces:
                    shape = landmark_predictor(gray, faces[0])
                    face_descriptor = face_rec_model.compute_face_descriptor(img, shape)
                    encodings.append(np.array(face_descriptor))
                    names.append(folder)
        return names, encodings

    def mark_attendance(self, identifier):
        today = datetime.now().strftime("%Y-%m-%d")
        month_name = datetime.now().strftime("%B")
        time_now = datetime.now().strftime("%H:%M:%S")

        try:
            roll, name, dept = identifier.split('_', 2)
        except:
            roll, name, dept = "Unknown", identifier, "Unknown"

        entry = {"Roll": roll, "Name": name, "Department": dept, "Date": today, "Time": time_now}

        if identifier in self.already_marked_today:
            self.feedback_label.config(text=f"{name} already marked today!", foreground="red")
            winsound.Beep(600, 200)
            return

        try:
            if os.path.exists(EXCEL_FILENAME):
                wb = openpyxl.load_workbook(EXCEL_FILENAME)
            else:
                wb = openpyxl.Workbook()

            if month_name not in wb.sheetnames:
                sheet = wb.create_sheet(month_name)
                sheet.append(["Roll", "Name", "Department", "Date", "Time"])
            else:
                sheet = wb[month_name]

            is_present_today = False
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == roll and row[3] == today:
                    is_present_today = True
                    break

            if not is_present_today:
                sheet.append([roll, name, dept, today, time_now])
                wb.save(EXCEL_FILENAME)
                print(f"[Excel] Attendance saved for {name} at {time_now}")

                file_exists = os.path.exists(CSV_FILENAME)
                with open(CSV_FILENAME, "a") as f:
                    if not file_exists:
                        f.write("Roll,Name,Department,Date,Time\n")
                    f.write(f"{roll},{name},{dept},{today},{time_now}\n")
                print(f"[CSV] Attendance saved for {name} at {time_now}")

                self.feedback_label.config(text=f"{name} marked present!", foreground="green")
                winsound.Beep(800, 300)
                self.already_marked_today.add(identifier)
            else:
                self.feedback_label.config(text=f"{name} already marked today!", foreground="red")
                winsound.Beep(600, 200)
                self.already_marked_today.add(identifier)

        except PermissionError:
            messagebox.showerror("File Error", "Make sure the Excel file is not open!")
        except Exception as e:
            messagebox.showerror("Error", f"Error in marking attendance: {e}")

    def recognize_and_mark(self):
        names, encodings = self.load_registered_faces()
        cap = cv2.VideoCapture(0)

        while True:
            ret, frame = cap.read()
            if not ret:
                break
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_detector(gray)
            for face in faces:
                shape = landmark_predictor(gray, face)
                face_encoding = face_rec_model.compute_face_descriptor(frame, shape)
                face_np = np.array(face_encoding)
                distances = np.linalg.norm(encodings - face_np, axis=1)
                if len(distances):
                    min_dist = np.min(distances)
                    if min_dist < 0.6:
                        match_index = np.argmin(distances)
                        name = names[match_index]
                        self.mark_attendance(name)
                        cv2.putText(frame, name, (face.left(), face.top()-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
                        cv2.rectangle(frame, (face.left(), face.top()), (face.right(), face.bottom()), (0, 255, 0), 2)
                    else:
                        cv2.putText(frame, "Unknown", (face.left(), face.top()-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)
                        cv2.rectangle(frame, (face.left(), face.top()), (face.right(), face.bottom()), (0, 0, 255), 2)
            cv2.imshow("Real-Time Recognition - Press Q to Exit", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        cap.release()
        cv2.destroyAllWindows()

    # def test_mark_attendance_mock(self):
    #     mock_identifiers = [
    #         "101_John_Doe_CSE",
    #         "102_Jane_Smith_IT",
    #         "103_Bob_Brown_ECE"
    #     ]
    #     for identifier in mock_identifiers:
    #         self.mark_attendance(identifier)

# Run GUI
if __name__ == '__main__':
    root = tk.Tk()
    app = FaceAttendanceApp(root)
    root.mainloop()
